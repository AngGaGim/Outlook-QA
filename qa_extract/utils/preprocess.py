from bs4 import BeautifulSoup
import html
import re
import pandas as pd
from openpyxl.reader.excel import load_workbook


# 邮件内容清洗
def clean_content(html_content):
    # 去除HTML标签
    soup = BeautifulSoup(html_content, 'html.parser')
    text = soup.get_text(separator='\n')

    # 处理特殊编码（如&amp;nbsp;）
    text = html.unescape(text)

    # 过滤广告内容（如Pinterest邮件）
    if "pinterest.com" in text:
        text = text.split("条款和条件").strip()

    return text

# qa提取规则
def extract_qa_pairs(email):
    qa_pairs = []
    # {'@odata.etag': 'W/"CQAAABYAAAAdSF5E9j+rT5IzKKZHP871AACUlCXL"', 'id': 'AQMkADAwATNiZmYAZS04MGYwLThiZGItMDACLTAwCgBGAAADM6nnryUVYEyzPUKjFXCMbQcAHUheRPY-q0_SMyimRz-O9QAAAgEMAAAAHUheRPY-q0_SMyimRz-O9QAAAJSVHTMAAAA=', 'createdDateTime': '2025-02-27T10:10:38Z', 'lastModifiedDateTime': '2025-02-27T10:16:40Z', 'changeKey': 'CQAAABYAAAAdSF5E9j+rT5IzKKZHP871AACUlCXL', 'categories': [], 'receivedDateTime': '2025-02-27T10:10:38Z', 'sentDateTime': '2025-02-27T10:09:31Z', 'hasAttachments': False, 'internetMessageId': '<tencent_D206E3CFCABA088365944B7E4675731A3D06@qq.com>', 'subject': '12121212', 'bodyPreview': '1212121212dededdddd', 'importance': 'normal', 'parentFolderId': 'AQMkADAwATNiZmYAZS04MGYwLThiZGItMDACLTAwCgAuAAADM6nnryUVYEyzPUKjFXCMbQEAHUheRPY-q0_SMyimRz-O9QAAAgEMAAAA', 'conversationId': 'AQQkADAwATNiZmYAZS04MGYwLThiZGItMDACLTAwCgAQAEP5HH4fGfNAurABd_z_aVM=', 'conversationIndex': 'AQHbiP/ZQ/kcfh8Z80C6sAF37P5pUw==', 'isDeliveryReceiptRequested': None, 'isReadReceiptRequested': False, 'isRead': True, 'isDraft': False, 'webLink': 'https://outlook.live.com/owa/?ItemID=AQMkADAwATNiZmYAZS04MGYwLThiZGItMDACLTAwCgBGAAADM6nnryUVYEyzPUKjFXCMbQcAHUheRPY%2Fq0%2BSMyimRz%2FO9QAAAgEMAAAAHUheRPY%2Fq0%2BSMyimRz%2FO9QAAAJSVHTMAAAA%3D&exvsurl=1&viewmodel=ReadMessageItem', 'inferenceClassification': 'other', 'body': {'contentType': 'html', 'content': '<html><head>\r\n<meta http-equiv="Content-Type" content="text/html; charset=utf-8"></head><body><div>1212121212dededdddd</div></body></html>'}, 'sender': {'emailAddress': {'name': '________honey_______', 'address': 'hjy145@foxmail.com'}}, 'from': {'emailAddress': {'name': '________honey_______', 'address': 'hjy145@foxmail.com'}}, 'toRecipients': [{'emailAddress': {'name': 'q&nbsp;q', 'address': 'wsjda5s@outlook.com'}}], 'ccRecipients': [], 'bccRecipients': [], 'replyTo': [], 'flag': {'flagStatus': 'notFlagged'}}
    # 主对话提取（参考）
    body = clean_content(email["body"]["content"])
    subject = email.get("subject",None)

    if subject and any(keyword in subject for keyword in ["回复", "Re", "RE"]):
        # 分割问题与回答（示例数据中发件人变化标识对话）
        q_part = email["bodyPreview"].split("________________________________")
        a_part = body.split("________________________________")
        qa_pairs.append(("Q", q_part))
        qa_pairs.append(("A", a_part))
    else:
        qa_pairs.append(("Q", body))  # 初始问题

    return qa_pairs



# 增强版字符清洗（参考）
def clean_non_printable(text):
    if pd.isnull(text):
        return text
    text = str(text)

    # 保留换行符，清洗其他控制字符和特殊字符
    text = re.sub(r'[\x00-\x1F\x7F\u2028\u2029]', '[111111]', text)  # 移除控制字符
    text = re.sub(r'\[111111\]+', '\n', text)
    text = re.sub(r'\n+', '\n', text)

    return text


def read_csv(csv_file_path):
    df = pd.read_csv(csv_file_path, dtype=str)
    # body_preview = df.loc[1, 'content']
    # print(df.columns)

    for col in df.columns:
        if col == 'bodyPreview' or col == 'content':
            df[col] = df[col].apply(clean_non_printable)

    return df

def to_excel(df,sheet_name):
    # 读取现有的 Excel 文件
    file_path =r"/Users/txmmy/my-python-projects/outlook-qa/data/emails.xlsx"
    wb = load_workbook(file_path)

    # 检查 sheet 是否存在
    if sheet_name in wb.sheetnames:
        print(f"Sheet {sheet_name}  already exists.")
    else:
        # 添加新 sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            print("New sheet added to the Excel file.")

    # with (pd.ExcelWriter('/Users/txmmy/my-python-projects/outlook-qa/data/emails.xlsx',
    #                     engine='openpyxl',
    #                     # engine_kwargs={'options': {'strings_to_urls': False}}
    #       )
    #       as writer):
    #     df.to_excel(writer, index=False,sheet_name=sheet_name)
    # print("Excel file created.")

def remove_blockquote(body_html):
    soup = BeautifulSoup(body_html, 'html.parser')
    # 删除所有 <blockquote> 元素
    for blockquote in soup.find_all('blockquote'):
        print(f"find blockquote")
        blockquote.decompose()  # 删除 <blockquote> 元素及其子元素
    # 删除包含 <table> 元素的所有父级标签
    for table in soup.find_all('table'):
        print("find table")
        table.decompose()  # 删除包含的 <table> 元素及其子元素

    return str(soup)